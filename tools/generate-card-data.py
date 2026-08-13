#!/usr/bin/env python3
"""Generate trickcal-manager/cards.js from trickcal_datasheet.xlsx card sheets."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


CARD_EFFECT_SHEET = "カード効果"
CARD_SPECIAL_SHEET = "カード特殊効果"
KEY_MAP_CANDIDATES = (
    "card-effect-key-map.tsv",
    "card-effect-key-map - card-effect-key-map.tsv",
)

KIND_MAP = {
    "遺物": "artifact",
    "スペル": "spell",
    "artifact": "artifact",
    "spell": "spell",
}

DMG_TYPE_MAP = {
    "物理": "phys",
    "物理ダメージ": "phys",
    "魔法": "mag",
    "魔法ダメージ": "mag",
}

INLINE_EFFECT_KEYS = {
    "HP": "hpP",
    "最大HP": "hpP",
    "最大HP増加": "hpP",
    "攻撃力": "atkP",
    "攻撃力増加": "atkP",
    "魔法攻撃力増加": "atkP",
    "物理攻撃力増加": "atkP",
    "防御力": "defP",
    "防御力増加": "defP",
    "攻撃速度": "hasteP",
    "攻撃速度増加": "hasteP",
    "会心率": "critRateP",
    "会心": "critRateP",
    "会心ダメージ量": "critDmgP",
    "会心DMG": "critDmgP",
    "会心DMG増加": "critDmgP",
    "被会心率減少": "critResP",
    "会心抵抗": "critResP",
    "会心抵抗増加": "critResP",
    "会心被ダメージ量減少": "critDmgResP",
    "被会心ダメージ量減少": "critDmgResP",
    "会心DMG抵抗": "critDmgResP",
    "会心DMG抵抗増加": "critDmgResP",
    "HP治癒量": "healingP",
    "着用者のHP治癒量": "healingP",
    "HP回復量": "hpRecoveryP",
    "与ダメージ増加": "addP",
    "与ダメージ上昇": "addP",
    "ダメージ増加": "addP",
    "被ダメージ減少": "takenDmgP",
    "敵防御力減少": "enemyDefDownP",
    "毎秒SP回復量増加": "spRegenP",
}

# Compatibility for source rows whose value class has not yet been migrated.
# Once the datasheet row is "固定値", the generated result remains identical.
SPECIAL_EFFECT_VALUE_CLASS_OVERRIDES = {
    "artifact_chalice_of_origins_e01": "固定値",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def is_blank(value: object) -> bool:
    return clean(value) == ""


def as_number(value: object) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = clean(value)
    if text == "":
        return None
    try:
        num = float(text)
    except ValueError:
        return None
    return int(num) if num.is_integer() else num


def as_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = clean(value).lower()
    return text in {"true", "1", "yes", "y", "on", "〇", "○", "あり"}


def sheet_to_objects(workbook, sheet_name: str) -> list[dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"{sheet_name} sheet was not found.")
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_index = next((i for i, row in enumerate(rows) if any(not is_blank(v) for v in row)), None)
    if header_index is None:
        return []
    raw_headers = rows[header_index]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, raw in enumerate(raw_headers):
        header = clean(raw) or f"col{index + 1}"
        seen[header] = seen.get(header, 0) + 1
        headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")
    objects: list[dict[str, object]] = []
    for row in rows[header_index + 1 :]:
        if not any(not is_blank(v) for v in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        objects.append(item)
    return objects


def load_key_map(path: Path) -> dict[str, str]:
    mapping = dict(INLINE_EFFECT_KEYS)
    if path.is_dir():
        for name in KEY_MAP_CANDIDATES:
            candidate = path / name
            if candidate.exists():
                path = candidate
                break
    if not path.exists():
        return mapping
    with path.open("r", encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file, delimiter="\t")
        for row in reader:
            key = clean(row.get("内部キー", ""))
            if not key or key in {"condition", "target", "ref", "unknown", "state", "seconds", "count", "value"}:
                continue
            names = [clean(row.get("表示名", ""))]
            names.extend(part.strip() for part in clean(row.get("別名", "")).split(";") if part.strip())
            for name in names:
                mapping[name] = key
    return mapping


def resolve_key(label: object, key_map: dict[str, str]) -> str:
    text = clean(label)
    return key_map.get(text, INLINE_EFFECT_KEYS.get(text, ""))


def resolve_special_effect_key(label: object, value_class: object, key_map: dict[str, str]) -> str:
    key = resolve_key(label, key_map)
    if clean(value_class) == "固定値":
        return {
            "spRecoveryP": "spRecovery",
            "spRegenP": "spRegen",
        }.get(key, key)
    return key


def star_values(row: dict[str, object], prefix: str) -> list[int | float | None]:
    return [as_number(row.get(f"{prefix}_Star{star}")) for star in range(1, 6)]


def make_bonus_array(pairs: list[tuple[str, list[int | float | None]]]) -> list[dict[str, int | float]]:
    bonuses: list[dict[str, int | float]] = []
    for index in range(5):
        entry: dict[str, int | float] = {}
        for key, values in pairs:
            if not key or index >= len(values):
                continue
            value = values[index]
            if value is not None:
                entry[key] = value
        bonuses.append(entry)
    return bonuses


def has_any_bonus(bonuses: list[dict[str, int | float]]) -> bool:
    return any(bool(item) for item in bonuses)


def has_any_nonzero(values: list[int | float | None]) -> bool:
    return any(value not in (None, 0, 0.0) for value in values)


def unique_nonblank(values: list[object]) -> list[object]:
    result: list[object] = []
    for value in values:
        if value is None or value == "":
            continue
        if value not in result:
            result.append(value)
    return result


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def build_id_aliases(
    base_rows: list[dict[str, object]],
    special_rows: list[dict[str, object]],
    card_map_path: Path,
    effect_map_path: Path,
) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    card_rows = read_tsv(card_map_path)
    effect_rows = read_tsv(effect_map_path)
    current_card_ids = {clean(row.get("id")) for row in base_rows if clean(row.get("id"))}
    mapped_card_ids = {clean(row.get("newId")) for row in card_rows if clean(row.get("newId"))}
    extra_cards = sorted(mapped_card_ids - current_card_ids)
    if extra_cards:
        raise ValueError(f"card id replacement target is missing from datasheet: {extra_cards}")

    current_effect_ids = {
        (clean(row.get("id")), clean(row.get("effectId")))
        for row in special_rows
        if clean(row.get("id")) and clean(row.get("effectId"))
    }
    mapped_effect_ids = {
        (clean(row.get("newCardId")), clean(row.get("newEffectId")))
        for row in effect_rows
        if clean(row.get("newCardId")) and clean(row.get("newEffectId"))
    }
    extra_effects = sorted(mapped_effect_ids - current_effect_ids)
    if extra_effects:
        raise ValueError(f"card effect replacement target is missing from datasheet: {extra_effects}")

    card_aliases = {
        clean(row.get("oldId")): clean(row.get("newId"))
        for row in card_rows
        if clean(row.get("oldId"))
        and clean(row.get("newId"))
        and clean(row.get("oldId")) != clean(row.get("newId"))
    }
    effect_aliases = {
        f"{clean(row.get('oldCardId'))}|{clean(row.get('oldEffectId'))}": {
            "cardId": clean(row.get("newCardId")),
            "effectId": clean(row.get("newEffectId")),
        }
        for row in effect_rows
        if clean(row.get("oldCardId"))
        and clean(row.get("oldEffectId"))
        and (
            clean(row.get("oldCardId")) != clean(row.get("newCardId"))
            or clean(row.get("oldEffectId")) != clean(row.get("newEffectId"))
        )
    }
    return card_aliases, effect_aliases


def build_cards(base_rows: list[dict[str, object]], special_rows: list[dict[str, object]], key_map: dict[str, str]):
    cards_by_id: dict[str, dict[str, object]] = {}
    solder_data: dict[str, dict[int, dict[str, int | float]]] = {}

    for row in base_rows:
        card_id = clean(row.get("id"))
        if not card_id:
            continue
        kind = KIND_MAP.get(clean(row.get("種別")), clean(row.get("種別")) or "artifact")
        costs = [as_number(row.get(f"Cost_Star{star}")) for star in range(1, 6)]
        cost_values = unique_nonblank(costs)
        card: dict[str, object] = {
            "id": card_id,
            "kind": kind,
            "rarity": clean(row.get("レア度")),
            "name": clean(row.get("カード名")),
        }
        if as_bool(row.get("愛用")):
            card["signature"] = True
        favorite = clean(row.get("愛用使徒"))
        if favorite:
            card["favoriteCharacter"] = favorite
        if cost_values:
            card["cost"] = cost_values[0]
            if len(cost_values) > 1 or any(value is None for value in costs):
                card["costByStar"] = [value or 0 for value in costs]

        bonus_pairs = []
        for slot in ("A", "B"):
            key = resolve_key(row.get(f"効果{slot}"), key_map)
            values = star_values(row, f"効果{slot}")
            if key and has_any_nonzero(values):
                bonus_pairs.append((key, values))
        bonuses = make_bonus_array(bonus_pairs)
        if has_any_bonus(bonuses):
            card["bonusesByStar"] = bonuses

        solder_pairs_by_level = {1: [], 2: []}
        for slot in ("A", "B"):
            key = resolve_key(row.get(f"はんだ効果{slot}"), key_map)
            if not key:
                continue
            for level in (1, 2):
                value = as_number(row.get(f"はんだ効果値{slot}+{level}"))
                if value is not None:
                    solder_pairs_by_level[level].append((key, value))
        solder_levels: dict[int, dict[str, int | float]] = {}
        for level, pairs in solder_pairs_by_level.items():
            entry: dict[str, int | float] = {}
            for key, value in pairs:
                entry[key] = value
            if entry:
                solder_levels[level] = entry
        if solder_levels:
            solder_data[card_id] = solder_levels

        card["conditionalEffects"] = []
        cards_by_id[card_id] = card

    for row in special_rows:
        card_id = clean(row.get("id"))
        if not card_id or card_id not in cards_by_id:
            continue
        effect_id = clean(row.get("effectId")) or make_effect_id(row)
        effect_label = clean(row.get("効果")) or effect_id
        condition = clean(row.get("条件1"))
        condition2 = clean(row.get("条件2"))
        target = clean(row.get("効果対象"))
        attack_type = clean(row.get("攻撃タイプ"))
        duration = clean(row.get("持続時間"))
        reference = clean(row.get("参照"))
        value_kind = SPECIAL_EFFECT_VALUE_CLASS_OVERRIDES.get(
            effect_id,
            clean(row.get("値分類")),
        )
        reset_condition = clean(row.get("リセット条件"))
        key = resolve_special_effect_key(effect_label, value_kind, key_map)
        values = [as_number(row.get(f"特殊効果_Star{star}")) for star in range(1, 6)]
        bonuses = make_bonus_array([(key, values)]) if key and any(value is not None for value in values) else []

        label_parts = [part for part in [condition, condition2, attack_type, effect_label] if part]
        label = " ".join(label_parts) if label_parts else effect_label
        short_label = " ".join(part for part in [attack_type, effect_label] if part) or effect_label
        descriptions = build_descriptions(row, values, key, value_kind)
        effect: dict[str, object] = {
            "id": effect_id,
            "type": "toggle" if bonuses else "info",
            "label": label,
        }
        if short_label and short_label != label:
            effect["shortLabel"] = short_label
        if value_kind:
            effect["valueClass"] = value_kind
        if duration:
            effect["duration"] = duration
        if duration:
            effect["duration"] = duration
        # 「同効果」と「同一使徒」はランタイムで判定範囲が異なるため、
        # ここで単一の nonStacking フラグへ畳み込まない。
        if as_bool(row.get("同効果非スタック")):
            effect["nonStackingSameEffect"] = True
        if as_bool(row.get("同一使徒非スタック")):
            effect["nonStackingSameApostle"] = True
        if attack_type in DMG_TYPE_MAP:
            effect["onlyWhenDmgType"] = DMG_TYPE_MAP[attack_type]
        meta = [part for part in [target, duration and f"持続:{duration}", reference and f"参照:{reference}", value_kind, reset_condition and f"リセット:{reset_condition}"] if part]
        if meta:
            effect["description"] = " / ".join(meta)
        if descriptions:
            effect["descriptionByStar"] = descriptions
        if bonuses:
            effect["bonusesByStar"] = bonuses
        cards_by_id[card_id].setdefault("conditionalEffects", []).append(effect)

    artifacts: list[dict[str, object]] = []
    spells: list[dict[str, object]] = []
    for row in base_rows:
        card_id = clean(row.get("id"))
        card = cards_by_id.get(card_id)
        if not card:
            continue
        if not card.get("conditionalEffects"):
            card.pop("conditionalEffects", None)
        if card.get("kind") == "spell":
            spells.append(card)
        else:
            artifacts.append(card)
    return {"artifacts": artifacts, "spells": spells}, solder_data


def make_effect_id(row: dict[str, object]) -> str:
    source = "_".join(clean(row.get(key)) for key in ("条件1", "攻撃タイプ", "効果") if clean(row.get(key)))
    source = re.sub(r"\W+", "_", source, flags=re.ASCII).strip("_")
    return source or "effect"


def build_descriptions(row: dict[str, object], values: list[int | float | None], key: str, value_class: str = "") -> list[str]:
    effect_label = clean(row.get("効果"))
    condition = clean(row.get("条件1"))
    target = clean(row.get("効果対象"))
    if not any(value is not None for value in values):
        return []
    descriptions: list[str] = []
    for value in values:
        if value is None:
            descriptions.append("")
            continue
        suffix = "%" if value_class == "倍率" else ""
        value_text = clean(value)
        has_embedded_value = bool(re.search(
            rf"(?<!\d)[+-]?{re.escape(value_text)}(?:%|秒|回)?$",
            effect_label,
        ))
        text = effect_label if has_embedded_value else f"{effect_label}{value_text}{suffix}"
        details = [part for part in [condition, target] if part]
        if details:
            text += f" ({' / '.join(details)})"
        descriptions.append(text)
    return descriptions


def stringify_js(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, indent=4)


def render_cards_js(
    card_library: dict[str, list[dict[str, object]]],
    solder_data: dict[str, dict[int, dict[str, int | float]]],
    card_id_aliases: dict[str, str],
    card_effect_id_aliases: dict[str, dict[str, str]],
) -> str:
    return (
        "// Trickcal Damage Calculator - Card Data (generated from datasheet)\n"
        "// Edit tools/trickcal_datasheet.xlsx and rebuild with tools/generate-card-data.py.\n\n"
        f"const CARD_LIBRARY = {stringify_js(card_library)};\n\n"
        f"const CARD_SOLDER_DATA = {stringify_js(solder_data)};\n\n"
        f"const CARD_ID_ALIASES = {stringify_js(card_id_aliases)};\n\n"
        f"const CARD_EFFECT_ID_ALIASES = {stringify_js(card_effect_id_aliases)};\n\n"
        "function resolveCardIdAlias(id) {\n"
        "    const key = String(id || '');\n"
        "    return CARD_ID_ALIASES[key] || key;\n"
        "}\n\n"
        "function migrateCardStateMap(source) {\n"
        "    const input = source && typeof source === 'object' ? source : {};\n"
        "    const result = {};\n"
        "    for (const [id, value] of Object.entries(input)) {\n"
        "        if (!CARD_ID_ALIASES[id]) result[id] = value;\n"
        "    }\n"
        "    for (const [id, value] of Object.entries(input)) {\n"
        "        const currentId = resolveCardIdAlias(id);\n"
        "        if (!Object.prototype.hasOwnProperty.call(result, currentId)) result[currentId] = value;\n"
        "    }\n"
        "    return result;\n"
        "}\n\n"
        "function migrateCardEffectStateMap(source) {\n"
        "    const input = source && typeof source === 'object' ? source : {};\n"
        "    const result = {};\n"
        "    for (const [rawKey, value] of Object.entries(input)) {\n"
        "        let key = String(rawKey || '');\n"
        "        for (const [legacyKey, current] of Object.entries(CARD_EFFECT_ID_ALIASES)) {\n"
        "            const separator = legacyKey.indexOf('|');\n"
        "            const oldCardId = legacyKey.slice(0, separator);\n"
        "            const oldEffectId = legacyKey.slice(separator + 1);\n"
        "            const cardToken = `:${oldCardId}:`;\n"
        "            const effectToken = `:${oldEffectId}:`;\n"
        "            if (!key.includes(cardToken) || !key.includes(effectToken)) continue;\n"
        "            key = key.replace(cardToken, `:${current.cardId}:`);\n"
        "            key = key.replace(effectToken, `:${current.effectId}:`);\n"
        "            break;\n"
        "        }\n"
        "        result[key] = value;\n"
        "    }\n"
        "    return result;\n"
        "}\n\n"
        "for (const card of [...CARD_LIBRARY.artifacts, ...CARD_LIBRARY.spells]) {\n"
        "    if (CARD_SOLDER_DATA[card.id]) {\n"
        "        card.solderBonuses = CARD_SOLDER_DATA[card.id];\n"
        "    }\n"
        "}\n\n"
        "const CARD_INDEX = Object.fromEntries(\n"
        "    [...CARD_LIBRARY.artifacts, ...CARD_LIBRARY.spells].map(card => [card.id, card])\n"
        ");\n"
    )


def generate(
    input_path: Path,
    output_path: Path,
    key_map_path: Path,
    card_map_path: Path,
    effect_map_path: Path,
) -> None:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    key_map = load_key_map(key_map_path)
    base_rows = sheet_to_objects(workbook, CARD_EFFECT_SHEET)
    special_rows = sheet_to_objects(workbook, CARD_SPECIAL_SHEET)
    card_library, solder_data = build_cards(base_rows, special_rows, key_map)
    card_id_aliases, card_effect_id_aliases = build_id_aliases(
        base_rows, special_rows, card_map_path, effect_map_path
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        render_cards_js(card_library, solder_data, card_id_aliases, card_effect_id_aliases),
        encoding="utf-8",
    )
    print(f"Built {output_path} from {input_path}")
    print(f"Cards: artifacts={len(card_library['artifacts'])}, spells={len(card_library['spells'])}, solder={len(solder_data)}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="trickcal_datasheet.xlsx", help="Path to trickcal_datasheet.xlsx")
    parser.add_argument("--output", default="../cards.js", help="Path to generated cards.js")
    parser.add_argument("--key-map", default="card-effect-key-map.tsv", help="Path to card effect key map TSV")
    parser.add_argument("--card-id-map", default="replacement/cardId.tsv", help="Path to card ID replacement TSV")
    parser.add_argument("--card-effect-id-map", default="replacement/cardEffectId.tsv", help="Path to card effect ID replacement TSV")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    input_path = Path(args.input)
    output_path = Path(args.output)
    key_map_path = Path(args.key_map)
    card_map_path = Path(args.card_id_map)
    effect_map_path = Path(args.card_effect_id_map)
    if not input_path.exists():
        input_path = script_dir / args.input
    if not output_path.is_absolute():
        output_path = script_dir / output_path
    if not key_map_path.exists():
        key_map_path = script_dir / args.key_map
    if not card_map_path.exists():
        card_map_path = script_dir / args.card_id_map
    if not effect_map_path.exists():
        effect_map_path = script_dir / args.card_effect_id_map
    generate(
        input_path.resolve(),
        output_path.resolve(),
        key_map_path.resolve(),
        card_map_path.resolve(),
        effect_map_path.resolve(),
    )


if __name__ == "__main__":
    main()
