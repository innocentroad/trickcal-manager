import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


SHEETS = {
    "basic": "使徒基礎設定",
    "skills": "スキル",
    "favoriteCard": "愛用カード",
    "asideStats": "アサイド ステ効果",
    "asideSpecials": "アサイド 特殊効果",
    "board": "ボード設定",
}


def normalize_cell(value):
    if value is None:
        return ""
    return value


def export_sheet(workbook, sheet_name, output_path):
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"sheet not found: {sheet_name}")

    worksheet = workbook[sheet_name]
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        for row in worksheet.iter_rows(values_only=True):
            values = [normalize_cell(value) for value in row]
            if not any(str(value).strip() for value in values):
                continue
            writer.writerow(values)


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: export-apostle-csvs.py <trickcal_datasheet.xlsx> <out_dir>")

    xlsx_path = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    for out_name, sheet_name in SHEETS.items():
        export_sheet(workbook, sheet_name, out_dir / f"{out_name}.csv")


if __name__ == "__main__":
    main()
