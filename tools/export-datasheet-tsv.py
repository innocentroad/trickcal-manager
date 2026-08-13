#!/usr/bin/env python3
"""Export every worksheet in the Trickcal datasheet to UTF-8 TSV files."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def filename_for_sheet(index: int, sheet_name: str) -> str:
    safe_name = INVALID_FILENAME_CHARS.sub("_", sheet_name).strip(" .") or "sheet"
    return f"{index:02d}_{safe_name}.tsv"


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def export_workbook(input_path: Path, output_dir: Path) -> dict[str, Any]:
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    manifest: dict[str, Any] = {
        "source": str(input_path.resolve()),
        "sheets": [],
    }
    try:
        for index, worksheet in enumerate(workbook.worksheets, start=1):
            filename = filename_for_sheet(index, worksheet.title)
            output_path = output_dir / filename
            row_count = 0
            column_count = 0
            with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
                for row in worksheet.iter_rows(values_only=True):
                    values = [cell_text(value) for value in row]
                    while values and values[-1] == "":
                        values.pop()
                    writer.writerow(values)
                    row_count += 1
                    column_count = max(column_count, len(values))
            manifest["sheets"].append(
                {
                    "name": worksheet.title,
                    "file": filename,
                    "rows": row_count,
                    "columns": column_count,
                }
            )
    finally:
        workbook.close()

    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export Trickcal datasheet worksheets as analysis-friendly TSV files."
    )
    parser.add_argument("--input", default="tools/trickcal_datasheet.xlsx")
    parser.add_argument("--output-dir", default="tmp/datasheet-tsv")
    args = parser.parse_args()

    manifest = export_workbook(Path(args.input), Path(args.output_dir))
    print(f"Exported {len(manifest['sheets'])} sheets to {args.output_dir}")


if __name__ == "__main__":
    main()
