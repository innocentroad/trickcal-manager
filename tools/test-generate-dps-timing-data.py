#!/usr/bin/env python3
"""Focused checks for the skillmotion DPS component-support parser."""
from __future__ import annotations

import importlib.util
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent
SPEC = importlib.util.spec_from_file_location("dps_timing_generator", ROOT / "generate-dps-timing-data.py")
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
SPEC.loader.exec_module(MODULE)


def make_book(rows):
    book = Workbook()
    sheet = book.active
    sheet.title = "対応状況"
    sheet.append(["id", "使徒名", "通常", "アサイド", "愛用品", "備考"])
    for row in rows:
        sheet.append(row)
    return book


with tempfile.TemporaryDirectory() as directory:
    path = Path(directory) / "fixture.xlsx"
    make_book([
        ["Chloe", "クロエ", "暫定", "", "済", "表示確認"],
        ["Sylla", "シーラ", "済", "途中", "未", None],
        ["Epica", "エピカ", "未", None, None, None],
        ["Future", "将来", "途中", "済", "暫定", None],
    ]).save(path)
    statuses = MODULE.support_status_rows(load_workbook(path), "対応状況")
    assert statuses["chloe"]["statuses"] == {"normal": "暫定", "aside": "未", "favorite": "済"}
    assert statuses["chloe"]["note"] == "表示確認"
    assert {item["statuses"]["normal"] for item in statuses.values()} == {"済", "暫定", "途中", "未"}
    assert statuses["epica"]["statuses"]["aside"] == "未", "empty component must safely map to 未"
    assert statuses["epica"]["statuses"]["favorite"] == "未", "empty component must safely map to 未"

    make_book([["Chloe", "クロエ", "済", "済", "済", None], ["chloe", "クロエ", "暫定", "済", "済", None]]).save(path)
    try:
        MODULE.support_status_rows(load_workbook(path), "対応状況")
        raise AssertionError("duplicate normalized id must fail")
    except ValueError as error:
        assert "重複" in str(error)

    make_book([["", "クロエ", "済", "済", "済", None]]).save(path)
    try:
        MODULE.support_status_rows(load_workbook(path), "対応状況")
        raise AssertionError("empty id must fail")
    except ValueError as error:
        assert "idが空" in str(error)

    make_book([["Chloe", "クロエ", "不明", "済", "済", None]]).save(path)
    try:
        MODULE.support_status_rows(load_workbook(path), "対応状況")
        raise AssertionError("unknown component status must fail")
    except ValueError as error:
        assert "未知" in str(error)

workbook_path = ROOT / "trickcal_skillmotion.xlsx"
data = MODULE.build_data(workbook_path, dict(MODULE.DEFAULT_SHEETS))
assert data["version"] == 8
assert data["supportStatuses"]["chloe"]["statuses"].keys() == {"normal", "aside", "favorite"}
assert data["apostles"]["chloe"]["implementationStatuses"].keys() == {"normal", "aside", "favorite"}
summary = data["summary"]["implementationStatuses"]
assert summary.keys() == {"normal", "aside", "favorite"}
assert all(set(counts) == {"済", "暫定", "途中", "未"} for counts in summary.values())
assert not any("Renewa_enhanced_e01" in warning for warning in data["warnings"]), "Barong source correction should remove cross-ID warning"

warnings: list[str] = []
assert MODULE.normalize_cross_apostle_effect_id(
    "Barong", "enhancedAttack", "Renewa_enhanced_e01", warnings, "fixture 1行"
) == "Barong_enhanced_e01"
assert len(warnings) == 1
assert MODULE.normalize_cross_apostle_effect_id(
    "Barong", "enhancedAttack", "artifact_barong_cursed_doll_e01", warnings, "fixture 2行"
) == "artifact_barong_cursed_doll_e01"
assert MODULE.normalize_cross_apostle_effect_id(
    "Barong", "enhancedAttack", "artifact_some_enhanced_e01", warnings, "fixture 3行"
) == "artifact_some_enhanced_e01"
assert MODULE.normalize_cross_apostle_effect_id(
    "Barong", "enhancedAttack", "external_enhanced_buff_e01", warnings, "fixture 4行"
) == "external_enhanced_buff_e01"

print("DPS timing component-support generator tests passed")
