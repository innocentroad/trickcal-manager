#!/usr/bin/env python3
"""Validate structured runtime-effect columns in trickcal_datasheet.xlsx."""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EFFECT_SHEETS = (
    "スキル効果",
    "アサイド特殊効果",
    "愛用カード効果",
    "カード特殊効果",
    "教主の権能効果",
)

BASE_SHEET_HEADERS = {
    "スキル基礎設定": ("id", "skillId", "使徒名", "スキル種別", "スキル名", "説明", "スキル発動条件種別", "スキル発動条件値"),
    "アサイド基礎設定": ("id", "skillId", "使徒名", "SLv", "効果区分", "効果説明"),
    "愛用カード基礎設定": ("id", "skillId", "使徒名", "解放Lv", "効果名", "説明"),
    "教主の権能基礎設定": ("id", "権能名", "説明", "コスト", "クールタイム秒"),
    "カード効果": ("id", "種別", "カード名", "レア度"),
}

PARENT_LINKS = {
    "スキル効果": ("スキル基礎設定", "skillId", "skillId"),
    "アサイド特殊効果": ("アサイド基礎設定", "skillId", "skillId"),
    "愛用カード効果": ("愛用カード基礎設定", "skillId", "skillId"),
    "教主の権能効果": ("教主の権能基礎設定", "id", "id"),
    "カード特殊効果": ("カード効果", "id", "id"),
}

STRUCTURED_HEADERS = (
    "effectId",
    "効果処理グループID",
    "処理順",
    "発動条件種別",
    "発動条件値",
    "発動元ID",
    "適用条件種別",
    "適用条件値",
)

VALUE_OPTIONAL_TRIGGER_TYPES = {
    "カード選択時",
    "戦闘開始時",
    "ウェーブ開始時",
    "低学年スキル使用時",
    "低学年スキル発動時",
    "低学年スキル終了時",
    "低学年スキル命中時",
    "高学年スキル使用時",
    "高学年スキル発動時",
    "高学年スキル終了時",
    "高学年スキル命中時",
    "普通攻撃命中時",
    "基本攻撃命中時",
    "強化攻撃終了時",
    "強化攻撃命中時",
    "生成物消滅時",
    "生成物攻撃時",
    "固有状態付与時",
}

VALUE_OPTIONAL_CONDITION_TYPES = {
    "ウェーブ内初回",
    "追加対象存在",
}

EFFECT_ID_PATTERN = re.compile(r"^[A-Za-z0-9_]+_e\d+$")
REFERENCE_LIKE_PATTERN = re.compile(
    r"(?:最大HP|現在HP|攻撃力|防御力|会心|会心DMG|会心ダメージ|SP|対象数|スキルLv|ステータス)$"
)


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass(frozen=True)
class Issue:
    severity: str
    sheet: str
    row: int
    effect_id: str
    message: str


def read_rows(workbook: Any, sheet_name: str) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    if sheet_name not in workbook.sheetnames:
        return [], []
    values = list(workbook[sheet_name].iter_rows(values_only=True))
    if not values:
        return [], []
    headers = [text(value) for value in values[0]]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values_row in enumerate(values[1:], start=2):
        if not any(text(value) for value in values_row):
            continue
        item = {
            header: values_row[index] if index < len(values_row) else None
            for index, header in enumerate(headers)
            if header
        }
        # Formula-filled FALSE cells can extend far below the actual table.
        if not text(item.get("id")) and not text(item.get("effectId")):
            continue
        rows.append((row_number, item))
    return headers, rows


def validate(input_path: Path) -> tuple[list[Issue], dict[str, int]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    issues: list[Issue] = []
    stats = defaultdict(int)
    all_effect_ids: set[str] = set()
    source_references: list[tuple[str, int, str, str]] = []
    condition_references: list[tuple[str, int, str, str]] = []

    parent_ids: dict[str, set[str]] = {}
    for sheet_name, required_headers in BASE_SHEET_HEADERS.items():
        headers, rows = read_rows(workbook, sheet_name)
        if not headers:
            issues.append(Issue("ERROR", sheet_name, 1, "", "シートまたは見出しがありません"))
            parent_ids[sheet_name] = set()
            continue
        missing = [header for header in required_headers if header not in headers]
        if missing:
            issues.append(Issue("ERROR", sheet_name, 1, "", f"基礎シートの列が不足: {', '.join(missing)}"))
        id_header = "skillId" if "skillId" in required_headers else "id"
        parent_ids[sheet_name] = {text(row.get(id_header)) for _, row in rows if text(row.get(id_header))}

    sheet_rows: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for sheet_name in EFFECT_SHEETS:
        headers, rows = read_rows(workbook, sheet_name)
        sheet_rows[sheet_name] = rows
        if not headers:
            issues.append(Issue("ERROR", sheet_name, 1, "", "シートまたは見出しがありません"))
            continue
        missing = [header for header in STRUCTURED_HEADERS if header not in headers]
        # Card effects remain legacy-compatible until their sheet is migrated.
        if sheet_name != "カード特殊効果" and missing:
            issues.append(Issue("ERROR", sheet_name, 1, "", f"新書式の列が不足: {', '.join(missing)}"))
        elif sheet_name == "カード特殊効果" and missing:
            issues.append(Issue("INFO", sheet_name, 1, "", "カード特殊効果は旧書式互換で処理中"))

        seen_ids: dict[str, int] = {}
        seen_group_orders: dict[tuple[str, str], tuple[int, str]] = {}
        for row_number, row in rows:
            effect_id = text(row.get("effectId"))
            if not effect_id:
                issues.append(Issue("ERROR", sheet_name, row_number, "", "effectIdが空です"))
                continue
            stats["effects"] += 1
            all_effect_ids.add(effect_id)
            parent_sheet, effect_parent_key, _ = PARENT_LINKS[sheet_name]
            effect_parent_id = text(row.get(effect_parent_key))
            if effect_parent_id and effect_parent_id not in parent_ids.get(parent_sheet, set()):
                issues.append(Issue("ERROR", sheet_name, row_number, effect_id, f"基礎設定を解決できません: {parent_sheet}/{effect_parent_id}"))
                stats["orphanEffects"] += 1
            if effect_id in seen_ids:
                issues.append(Issue("ERROR", sheet_name, row_number, effect_id, f"effectId重複（先頭行: {seen_ids[effect_id]}）"))
            else:
                seen_ids[effect_id] = row_number

            group_id = text(row.get("効果処理グループID"))
            order = text(row.get("処理順"))
            if bool(group_id) != bool(order):
                issues.append(Issue("WARN", sheet_name, row_number, effect_id, "処理グループIDと処理順は両方設定してください"))
            if group_id and order:
                stats["grouped"] += 1
                key = (group_id, order)
                if key in seen_group_orders:
                    previous_row, previous_id = seen_group_orders[key]
                    issues.append(Issue("ERROR", sheet_name, row_number, effect_id, f"処理順重複: {group_id}/{order}（{previous_id}, 行{previous_row}）"))
                else:
                    seen_group_orders[key] = (row_number, effect_id)

            trigger_type = text(row.get("発動条件種別"))
            trigger_value = text(row.get("発動条件値"))
            trigger_source = text(row.get("発動元ID"))
            condition_type = text(row.get("適用条件種別"))
            condition_value = text(row.get("適用条件値"))
            attack_category = text(row.get("攻撃分類")) or text(row.get("攻撃タイプ"))
            legacy_condition = text(row.get("条件"))
            target_skill = text(row.get("対象スキル"))
            reference = text(row.get("参照"))
            if trigger_type:
                stats["triggered"] += 1
            if condition_type:
                stats["conditioned"] += 1
            if trigger_value and not trigger_type:
                issues.append(Issue("WARN", sheet_name, row_number, effect_id, "発動条件値がありますが発動条件種別が空です"))
            if trigger_source and not trigger_type:
                issues.append(Issue("WARN", sheet_name, row_number, effect_id, "発動元IDがありますが発動条件種別が空です"))
            if condition_value and not condition_type:
                issues.append(Issue("WARN", sheet_name, row_number, effect_id, "適用条件値がありますが適用条件種別が空です"))
            if trigger_type and not trigger_value and trigger_type not in VALUE_OPTIONAL_TRIGGER_TYPES:
                stats["triggerWithoutValue"] += 1
            if condition_type and not condition_value and condition_type not in VALUE_OPTIONAL_CONDITION_TYPES:
                issues.append(Issue("WARN", sheet_name, row_number, effect_id, f"適用条件値が空です: {condition_type}"))
            if legacy_condition and not trigger_type and not condition_type and not attack_category:
                issues.append(Issue("WARN", sheet_name, row_number, effect_id, f"旧条件のみ: {legacy_condition}"))
                stats["legacyOnly"] += 1
            if target_skill and not reference and REFERENCE_LIKE_PATTERN.search(target_skill):
                issues.append(Issue(
                    "WARN",
                    sheet_name,
                    row_number,
                    effect_id,
                    f"対象スキルに参照値らしい値があります（列ずれ候補）: {target_skill}",
                ))

            if EFFECT_ID_PATTERN.match(trigger_source):
                source_references.append((sheet_name, row_number, effect_id, trigger_source))
            if EFFECT_ID_PATTERN.match(condition_value):
                condition_references.append((sheet_name, row_number, effect_id, condition_value))

    for sheet_name, row_number, effect_id, reference in source_references:
        if reference not in all_effect_ids:
            issues.append(Issue("WARN", sheet_name, row_number, effect_id, f"発動元effectIdを解決できません: {reference}"))
    for sheet_name, row_number, effect_id, reference in condition_references:
        if reference not in all_effect_ids:
            issues.append(Issue("WARN", sheet_name, row_number, effect_id, f"適用条件effectIdを解決できません: {reference}"))

    stats["errors"] = sum(issue.severity == "ERROR" for issue in issues)
    stats["warnings"] = sum(issue.severity == "WARN" for issue in issues)
    stats["info"] = sum(issue.severity == "INFO" for issue in issues)
    return issues, dict(stats)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="trickcal_datasheet.xlsx")
    parser.add_argument("--strict", action="store_true", help="Exit non-zero when warnings remain.")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    input_path = Path(args.input)
    if not input_path.exists():
        input_path = script_dir / input_path
    issues, stats = validate(input_path.resolve())
    for issue in issues:
        location = f"{issue.sheet}:{issue.row}"
        suffix = f" [{issue.effect_id}]" if issue.effect_id else ""
        print(f"{issue.severity}: {location}{suffix} {issue.message}")
    print(
        "SUMMARY: "
        + " ".join(f"{key}={stats.get(key, 0)}" for key in (
            "effects", "grouped", "triggered", "conditioned", "legacyOnly", "orphanEffects", "errors", "warnings", "info"
        ))
    )
    if stats.get("errors", 0) or (args.strict and stats.get("warnings", 0)):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
