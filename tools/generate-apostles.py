#!/usr/bin/env python3
"""Generate trickcal-manager/apostles.js from trickcal_datasheet.xlsx."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEETS = {
    "basic": "使徒基礎設定",
    "skillBasics": "スキル基礎設定",
    "skillEffects": "スキル効果",
    "favoriteCardBasics": "愛用カード基礎設定",
    "favoriteCardEffects": "愛用カード効果",
    "asideBasics": "アサイド基礎設定",
    "asideStats": "アサイド全体補正",
    "asideSpecials": "アサイド特殊効果",
    "uniqueStates": "固有状態基礎設定",
    "board": "ボード設定",
}

KEY_MAP = {
    "id": "id",
    "使徒名": "name",
    "(en)": "id",
    "en": "id",
    "レア度": "rarity",
    "エルダイン": "eldain",
    "性格": "personality",
    "種族": "race",
    "役割": "role",
    "配列": "position",
    "配置列": "position",
    "攻撃タイプ": "attackType",
    "攻撃Type": "attackType",
    "HPタイプ": "hpType",
    "HPTier": "hpType",
    "物理攻撃力タイプ": "atkPType",
    "物理攻撃力Tier": "atkPType",
    "魔法攻撃力タイプ": "atkMType",
    "魔法攻撃力Tier": "atkMType",
    "物理防御力タイプ": "defPType",
    "物理防御力Tier": "defPType",
    "魔法防御力タイプ": "defMType",
    "魔法防御力Tier": "defMType",
    "会心タイプ": "critType",
    "会心Tier": "critType",
    "会心DMGタイプ": "critDmgType",
    "会心DMGTier": "critDmgType",
    "会心抵抗タイプ": "critResType",
    "会心抵抗Tier": "critResType",
    "会心DMG抵抗タイプ": "critDmgResType",
    "会心DMG抵抗Tier": "critDmgResType",
    "SP回復": "initialSp",
    "初期SP": "initialSp",
    "毎秒SP回復量": "spRecoveryPerSecond",
    "戦闘力補正値A": "combatPowerCorrectionA",
    "戦闘力補正値B": "combatPowerCorrectionB",
    "戦闘力補正": "combatPowerCorrectionB",
    "weight_value_a": "combatPowerCorrectionB",
    "ボードタイプ": "boardType",
    "ボード形": "boardShape",
    "no": "no",
    "スキル種別": "skillType",
    "スキル名": "skillName",
    "効果名": "skillName",
    "説明": "description",
    "硬直秒": "stunSeconds",
    "クールタイム秒": "cooldownSeconds",
    "高学年クールタイム秒": "cooldownSeconds",
    "値の種類": "valueKind",
    "値分類": "valueClass",
    "効果タイプ": "effectType",
    "ダメージ補正区分": "damageModifierCategory",
    "効果スタック": "effectStack",
    "スタック数": "stackCount",
    "最大スタック": "maxStack",
    "最大スタック数": "maxStack",
    "最大スタック到達時消費": "consumeOnMaxStack",
    "スキル発動条件種別": "triggerType",
    "発動条件種別": "triggerType",
    "スキル発動条件値": "triggerValue",
    "発動条件値": "triggerValue",
    "効果処理グループID": "processGroupId",
    "スタックグループID": "stackGroupId",
    "スタック集約ID": "stackGroupId",
    "処理順": "processOrder",
    "発動元ID": "triggerSourceId",
    "適用条件種別": "conditionType",
    "適用条件値": "conditionValue",
    "発動条件": "condition",
    "条件": "condition",
    "効果対象": "effectTarget",
    "対象スキル": "targetSkill",
    "対象スキル名": "targetSkillName",
    "参照": "reference",
    "持続時間": "duration",
    "固定値": "fixedValue",
    "stateId": "stateId",
    "状態名": "name",
    "状態カテゴリ": "category",
    "値形式": "valueType",
    "管理単位": "scope",
    "初期値": "initialValue",
    "最小値": "minValue",
    "最大値": "maxValue",
    "増減単位": "step",
    "上限時処理": "capBehavior",
    "下限時処理": "floorBehavior",
    "解除区分": "dispelPolicy",
    "保持期間": "retention",
    "変化イベント基準": "changeEventBasis",
    "計算対応段階": "calculationSupportLevel",
    "検証状態": "verificationStatus",
    "備考": "notes",
    "出典URL": "sourceUrl",
    "遺物名": "cardName",
    "カード名": "cardName",
    "カード種別": "cardKind",
    "解放Lv": "unlockLevel",
    "Lv1_説明": "lv1Description",
    "効果1種別": "effect1Kind",
    "効果1値": "effect1Value",
    "効果2種別": "effect2Kind",
    "効果2値": "effect2Value",
    "アサイド名": "asideName",
    "Lv": "level",
    "SLv": "level",
    "Lv内名前": "levelName",
    "効果説明": "effectDescription",
    "効果区分": "effectCategory",
    "ステ適用": "statApplyTo",
    "ステータス適用": "statApplyTo",
    "ステ能力値": "statName",
    "上昇ステータス": "statName",
    "上昇%": "increaseP",
    "種族タイプ": "raceBoardType",
}

EFFECT_ATTACK_CATEGORY_HEADERS = {
    "攻撃タイプ",
    "攻撃分類",
    "攻撃分類上書き",
    "対象攻撃種別",
}

ALLOWED_VALUE_CLASSES = {
    "倍率",
    "与ダメージ量増加",
    "被ダメージ量減少",
    "固定値",
    "持続時間",
    "状態付与",
    "状態解除",
    "状態免疫",
    "免疫",
    "ヒット数",
    "対象数",
    "最大対象数",
    "回数",
    "被弾回数",
    "周期",
    "クールタイム",
    "条件",
    "SP量",
    "スキル変更",
    "最大スタック",
    "スタック数",
    "上限値",
    "解除",
    "召喚",
    "移動",
}

REFERENCE_VALUE_NAMES = {
    "最大HP", "対象の最大HP", "自身の最大HP",
    "攻撃力", "対象の攻撃力", "自身の攻撃力",
    "防御力", "対象の防御力", "自身の防御力",
}


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).replace("\ufeff", "").replace("\u2003", " ").replace("\u00a0", " ").strip()
    if text in {"", "-"}:
        return None
    if re.fullmatch(r"[+-]?\d+", text):
        # Excel の保存時に `1.0` が `1` へ正規化されても、生成済み
        # データの数値表記だけが全件変わらないよう従来の小数表記を保つ。
        # JavaScript 側では同じ Number であり、値の意味は変えない。
        return float(text)
    try:
        return float(text)
    except ValueError:
        return text


def has_value(value: Any) -> bool:
    return value is not None and value != ""


def normalize_id(value: Any) -> str | None:
    if not value:
        return None
    return re.sub(r"[^0-9a-z]+", "_", str(value).lower()).strip("_")


def key_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def read_sheet_rows(
    workbook: Any,
    sheet_name: str,
    *,
    effect_rows: bool = False,
    key_overrides: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"sheet not found: {sheet_name}")

    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    rows = [row for row in rows if any(str(cell).strip() for cell in row if cell is not None)]
    if not rows:
        return []

    headers = [str(cell or "").strip() for cell in rows[0]]
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        item: dict[str, Any] = {}
        for raw_key, raw_value in zip(headers, row):
            if not raw_key:
                continue
            key = (key_overrides or {}).get(raw_key)
            if key is None:
                key = (
                    "attackCategory"
                    if effect_rows and raw_key in EFFECT_ATTACK_CATEGORY_HEADERS
                    else KEY_MAP.get(raw_key, raw_key)
                )
            value = clean_value(raw_value)
            if key == "effectStack" and value is False:
                continue
            if has_value(value):
                item[key] = value
        if item.get("targetSkill") in REFERENCE_VALUE_NAMES and not item.get("reference"):
            item["reference"] = item.pop("targetSkill")
        if "id" in item:
            item["id"] = normalize_id(item["id"])
        items.append(item)
    return items


def compact_row(row: dict[str, Any], remove: tuple[str, ...] = ("id", "name")) -> dict[str, Any]:
    return {key: value for key, value in row.items() if key not in remove and has_value(value)}


def extract_levels(data: dict[str, Any]) -> dict[str, Any]:
    levels: dict[str, Any] = {}
    for index in range(1, 16):
        key = f"Lv{index}"
        if key in data:
            levels[str(index)] = data.pop(key)
    return levels


def has_effect_payload(data: dict[str, Any]) -> bool:
    keys = (
        "valueKind",
        "valueClass",
        "effectType",
        "attackCategory",
        "effectStack",
        "stackCount",
        "maxStack",
        "processGroupId",
        "stackGroupId",
        "processOrder",
        "triggerType",
        "triggerValue",
        "triggerSourceId",
        "conditionType",
        "conditionValue",
        "condition",
        "effectTarget",
        "targetSkill",
        "reference",
        "duration",
        "fixedValue",
        "levels",
    )
    return any(key in data and has_value(data[key]) for key in keys)


def apply_effect_defaults(data: dict[str, Any]) -> dict[str, Any]:
    if (
        "valueKind" in data
        and "valueClass" not in data
        and data.get("effectType") == "攻撃"
        and str(data["valueKind"]).endswith("ダメージ")
    ):
        data["valueClass"] = "倍率"
    return data


def add_effect_if_present(effects: list[dict[str, Any]], data: dict[str, Any]) -> list[dict[str, Any]]:
    data = apply_effect_defaults(data)
    if has_effect_payload(data):
        effects.append(data)
    return effects


def new_group_key(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    return "|".join("" if row.get(key) is None else key_text(row.get(key)) for key in keys)


def add_grouped_effect(
    items: list[dict[str, Any]],
    row: dict[str, Any],
    group_keys: tuple[str, ...],
    parent_keys: tuple[str, ...],
    effect_remove_keys: tuple[str, ...],
) -> list[dict[str, Any]]:
    group_key = new_group_key(row, group_keys)
    group = next((item for item in items if item.get("_groupKey") == group_key), None)
    if group is None:
        group = {"_groupKey": group_key, "effects": []}
        for key in parent_keys:
            value = row.get(key)
            if has_value(value):
                group[key] = value
        items.append(group)

    data = compact_row(row)
    for key in effect_remove_keys:
        data.pop(key, None)
    levels = extract_levels(data)
    if levels:
        data["levels"] = levels
    group["effects"] = add_effect_if_present(group["effects"], data)
    return items


def merge_effect_rows(
    basics: list[dict[str, Any]],
    effects: list[dict[str, Any]],
    warnings: list[str],
    source_name: str,
    include_unmatched_basics: bool = True,
) -> list[dict[str, Any]]:
    basics_by_skill_id = {
        str(row.get("skillId", "")): row
        for row in basics
        if has_value(row.get("skillId"))
    }
    merged: list[dict[str, Any]] = []
    used_skill_ids: set[str] = set()
    for effect in effects:
        skill_id = str(effect.get("skillId", ""))
        if not skill_id:
            continue
        original_skill_id = skill_id
        if source_name.startswith("aside ") and skill_id not in basics_by_skill_id:
            match = re.fullmatch(r"(.+?)_(\d+)_(global|battle)", skill_id)
            if match:
                candidate = f"{match.group(1)}_aside_{match.group(2)}_{match.group(3)}"
                if candidate in basics_by_skill_id:
                    skill_id = candidate
                    effect = {**effect, "skillId": candidate}
                    warnings.append(
                        f"{source_name} skillId normalized: {original_skill_id} -> {candidate}"
                    )
        basic = basics_by_skill_id.get(skill_id)
        if basic is None:
            warnings.append(f"{source_name} effect references missing skillId: {skill_id}")
            continue
        if basic.get("id") and effect.get("id") and basic["id"] != effect["id"]:
            warnings.append(
                f"{source_name} apostle id mismatch: {skill_id} / {basic['id']} != {effect['id']}"
            )
            continue
        merged.append({**basic, **effect})
        used_skill_ids.add(skill_id)
    if include_unmatched_basics:
        for skill_id, basic in basics_by_skill_id.items():
            if skill_id not in used_skill_ids:
                merged.append(dict(basic))
    return merged


def remove_internal_group_keys(items: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    if not items:
        return []
    for item in items:
        item.pop("_groupKey", None)
        if "effects" in item:
            item["effects"] = list(item["effects"])
    return items


def new_apostle_data(apostle_id: str, name: str, basic: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "id": apostle_id,
        "name": name,
        "basic": basic or {},
        "statTypes": {},
        "skills": [],
        "uniqueStates": [],
        "favoriteCard": {},
        "aside": {"levels": {}},
        "board": None,
    }


def ensure_apostle_from_row(
    by_id: dict[str, dict[str, Any]], row: dict[str, Any], warnings: list[str]
) -> dict[str, Any] | None:
    apostle_id = row.get("id")
    if not apostle_id:
        return None
    if apostle_id not in by_id:
        name = row.get("name") or apostle_id
        by_id[apostle_id] = new_apostle_data(apostle_id, name)
        warnings.append(f"basic row missing, created placeholder apostle: {apostle_id}")
    return by_id[apostle_id]


def ensure_aside_level(apostle: dict[str, Any], row: dict[str, Any]) -> dict[str, Any] | None:
    aside = apostle["aside"]
    if row.get("asideName") and "name" not in aside:
        aside["name"] = row["asideName"]

    level = row.get("level")
    if not level:
        return None

    level_key = key_text(level)
    if level_key not in aside["levels"]:
        aside["levels"][level_key] = {"name": row.get("levelName"), "stats": [], "effects": []}
    level_item = aside["levels"][level_key]
    if row.get("levelName") and not level_item.get("name"):
        level_item["name"] = row["levelName"]
    return level_item


def add_aside_stat(apostle: dict[str, Any], row: dict[str, Any]) -> None:
    level_item = ensure_aside_level(apostle, row)
    if level_item is None:
        return
    data = compact_row(row)
    for key in ("asideName", "level", "levelName", "effectDescription", "effectCategory"):
        data.pop(key, None)
    levels = extract_levels(data)
    if levels:
        data["levels"] = levels
    if data:
        level_item["stats"].append(data)


def add_aside_special(apostle: dict[str, Any], row: dict[str, Any]) -> None:
    level_item = ensure_aside_level(apostle, row)
    if level_item is None:
        return
    if row.get("effectDescription"):
        level_item["description"] = row["effectDescription"]
    data = compact_row(row)
    for key in ("asideName", "level", "levelName", "effectDescription", "effectCategory"):
        data.pop(key, None)
    levels = extract_levels(data)
    if levels:
        data["levels"] = levels
    level_item["effects"] = add_effect_if_present(level_item["effects"], data)


def ensure_favorite_card_level(apostle: dict[str, Any], row: dict[str, Any]) -> str:
    favorite = apostle["favoriteCard"]
    if row.get("cardName") and "name" not in favorite:
        favorite["name"] = row["cardName"]
    if row.get("cardKind") and "kind" not in favorite:
        favorite["kind"] = row["cardKind"]
    favorite.setdefault("levels", {})

    unlock_level = row.get("unlockLevel")
    if unlock_level is None:
        unlock_level = 0
    level_key = key_text(unlock_level)
    favorite["levels"].setdefault(level_key, [])
    return level_key


def add_favorite_card_effect(apostle: dict[str, Any], row: dict[str, Any]) -> None:
    level_key = ensure_favorite_card_level(apostle, row)
    groups = apostle["favoriteCard"]["levels"][level_key]
    group_key = new_group_key(row, ("skillId",))
    group = next((item for item in groups if item.get("_groupKey") == group_key), None)
    if group is None:
        group = {"_groupKey": group_key, "effects": []}
        if row.get("skillId"):
            group["skillId"] = row["skillId"]
        for key in ("skillName", "lv1Description"):
            value = row.get(key)
            if has_value(value):
                group["description" if key == "lv1Description" else key] = value
        groups.append(group)

    data = compact_row(row)
    for key in ("skillId", "cardName", "cardKind", "unlockLevel", "skillName", "lv1Description"):
        data.pop(key, None)
    levels = extract_levels(data)
    if levels:
        data["levels"] = levels
    group["effects"] = add_effect_if_present(group["effects"], data)


def build_library(workbook: Any) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], list[str]]:
    effect_sheets = {"skillEffects", "favoriteCardEffects", "asideSpecials"}
    rows = {
        key: read_sheet_rows(
            workbook,
            sheet_name,
            effect_rows=key in effect_sheets,
            key_overrides=(
                {"説明": "lv1Description"}
                if key == "favoriteCardBasics"
                else {
                    "スキル発動条件種別": "skillTriggerType",
                    "スキル発動条件値": "skillTriggerValue",
                }
                if key == "skillBasics"
                else None
            ),
        )
        for key, sheet_name in SHEETS.items()
    }
    warnings: list[str] = []
    rows["skills"] = merge_effect_rows(
        rows["skillBasics"], rows["skillEffects"], warnings, "skill"
    )
    rows["favoriteCard"] = merge_effect_rows(
        rows["favoriteCardBasics"], rows["favoriteCardEffects"], warnings, "favorite card"
    )
    rows["asideStats"] = merge_effect_rows(
        rows["asideBasics"], rows["asideStats"], warnings, "aside stat", False
    )
    rows["asideSpecials"] = merge_effect_rows(
        rows["asideBasics"], rows["asideSpecials"], warnings, "aside special", False
    )
    by_id: dict[str, dict[str, Any]] = {}

    for row in rows["basic"]:
        apostle_id = row.get("id")
        if not apostle_id:
            continue
        basic = compact_row(row)
        stat_types: dict[str, Any] = {}
        for out_key, in_key in (
            ("hp", "hpType"),
            ("atkP", "atkPType"),
            ("atkM", "atkMType"),
            ("defP", "defPType"),
            ("defM", "defMType"),
            ("crit", "critType"),
            ("critDmg", "critDmgType"),
            ("critRes", "critResType"),
            ("critDmgRes", "critDmgResType"),
        ):
            if in_key in basic:
                stat_types[out_key] = basic.pop(in_key)

        apostle = new_apostle_data(apostle_id, row.get("name", apostle_id), basic)
        apostle["statTypes"] = stat_types
        by_id[apostle_id] = apostle

    for row in rows["skills"]:
        apostle = ensure_apostle_from_row(by_id, row, warnings)
        if apostle is None:
            continue
        apostle["skills"] = add_grouped_effect(
            apostle["skills"],
            row,
            ("skillId",),
            ("skillId", "no", "skillType", "skillName", "description", "stunSeconds", "cooldownSeconds", "skillTriggerType", "skillTriggerValue"),
            ("skillId", "no", "skillType", "skillName", "description", "stunSeconds", "cooldownSeconds", "skillTriggerType", "skillTriggerValue"),
        )

    seen_unique_state_ids: set[str] = set()
    for row in rows["uniqueStates"]:
        state_id = row.get("stateId")
        owner_id = row.get("id")
        if not state_id:
            warnings.append("unique state definition missing stateId")
            continue
        if state_id in seen_unique_state_ids:
            warnings.append(f"duplicate unique state definition: {state_id}")
            continue
        seen_unique_state_ids.add(state_id)
        apostle = by_id.get(owner_id)
        if apostle is None:
            warnings.append(f"unique state owner is missing: {state_id} / {owner_id}")
            continue
        state = compact_row(row, remove=("id",))
        state["ownerId"] = owner_id
        apostle["uniqueStates"].append(state)

    for row in rows["favoriteCard"]:
        apostle = ensure_apostle_from_row(by_id, row, warnings)
        if apostle is not None:
            add_favorite_card_effect(apostle, row)

    for row in rows["asideStats"]:
        apostle = ensure_apostle_from_row(by_id, row, warnings)
        if apostle is not None:
            add_aside_stat(apostle, row)

    for row in rows["asideSpecials"]:
        apostle = ensure_apostle_from_row(by_id, row, warnings)
        if apostle is not None:
            add_aside_special(apostle, row)

    for row in rows["board"]:
        apostle = ensure_apostle_from_row(by_id, row, warnings)
        if apostle is None:
            continue
        board = compact_row(row)
        if "X_pos" in board or "Y_pos" in board or "マス_type" in board:
            continue
        cells: dict[str, Any] = {}
        for key in list(board.keys()):
            if re.fullmatch(r"\d+-\d+", key):
                cells[key] = board.pop(key)
        board["cells"] = cells
        apostle["board"] = board

    for apostle in by_id.values():
        if not apostle["uniqueStates"]:
            apostle.pop("uniqueStates")
        apostle["skills"] = remove_internal_group_keys(apostle["skills"])
        for skill in apostle["skills"]:
            if "skillTriggerType" in skill:
                skill["triggerType"] = skill.pop("skillTriggerType")
            if "skillTriggerValue" in skill:
                skill["triggerValue"] = skill.pop("skillTriggerValue")

        favorite = apostle["favoriteCard"]
        if "levels" in favorite:
            favorite["levels"] = {
                key: remove_internal_group_keys(favorite["levels"][key])
                for key in sorted(favorite["levels"], key=lambda value: int(value))
            }

        aside = apostle["aside"]
        sorted_aside: dict[str, Any] = {}
        if "name" in aside:
            sorted_aside["name"] = aside["name"]
        sorted_aside["levels"] = {
            key: {
                **aside["levels"][key],
                "stats": list(aside["levels"][key]["stats"]),
                "effects": list(aside["levels"][key]["effects"]),
            }
            for key in sorted(aside["levels"], key=lambda value: int(value))
        }
        apostle["aside"] = sorted_aside

    library = list(by_id.values())
    validate_library(library, warnings)
    return library, by_id, warnings


def validate_library(library: list[dict[str, Any]], warnings: list[str]) -> None:
    for apostle in library:
        for skill in apostle["skills"]:
            if not has_value(skill.get("skillName")) and not skill.get("effects"):
                warnings.append(f"empty skill row: {apostle['id']} / skill:{skill.get('skillType')}")
            validate_effects(apostle["id"], f"skill:{skill.get('skillType')}", skill.get("effects", []), warnings)

        favorite = apostle["favoriteCard"]
        if "levels" in favorite:
            for level_key, groups in favorite["levels"].items():
                for group in groups:
                    validate_effects(
                        apostle["id"], f"favoriteCard:Lv{level_key}", group.get("effects", []), warnings
                    )

        for level_key, level_item in apostle["aside"]["levels"].items():
            validate_effects(apostle["id"], f"aside:Lv{level_key}", level_item.get("effects", []), warnings)


def validate_effects(
    apostle_id: str, source: str, effects: list[dict[str, Any]], warnings: list[str]
) -> None:
    for effect in effects:
        if effect.get("valueKind") and not effect.get("valueClass"):
            warnings.append(f"valueClass missing: {apostle_id} / {source} / {effect.get('valueKind')}")
        if effect.get("valueClass") and effect["valueClass"] not in ALLOWED_VALUE_CLASSES:
            warnings.append(
                f"unknown valueClass '{effect['valueClass']}': {apostle_id} / {source} / {effect.get('valueKind')}"
            )


def write_apostles_js(output_path: Path, library: list[dict[str, Any]]) -> None:
    json_text = json.dumps(library, ensure_ascii=False, indent=2)
    js = (
        "// Trickcal Damage Calculator - Apostle Data\n"
        "// Generated from: トリッカル使徒データ Google Sheet\n\n"
        f"const APOSTLE_LIBRARY = {json_text};\n\n"
        "const APOSTLE_INDEX = Object.fromEntries(APOSTLE_LIBRARY.map(apostle => [apostle.id, apostle]));\n"
    )
    output_path.write_text(js, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default=str(script_dir / "trickcal_datasheet.xlsx"), help="Path to xlsx")
    parser.add_argument("--output", default=str(script_dir.parent / "apostles.js"), help="Output apostles.js path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    library, by_id, warnings = build_library(workbook)
    write_apostles_js(output_path, library)

    skill_count = sum(len(apostle["skills"]) for apostle in library)
    print(f"Generated {output_path}")
    print(f"apostles={len(library)} skills={skill_count}")
    vivi = by_id.get("vivi", {"skills": [], "aside": {"levels": {}}})
    print(f"vivi skills={len(vivi['skills'])} asideLevels={len(vivi['aside']['levels'])}")
    if warnings:
        print(f"validation warnings={len(warnings)}")
        for warning in warnings[:30]:
            print(f"WARN: {warning}")
        if len(warnings) > 30:
            print(f"WARN: ...and {len(warnings) - 30} more")


if __name__ == "__main__":
    main()
