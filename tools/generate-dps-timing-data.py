#!/usr/bin/env python3
"""Generate normalized DPS timing data from the skill-motion workbook.

The default source is the standalone ``trickcal_skillmotion.xlsx`` workbook.
The same sheet names can later be moved into ``trickcal_datasheet.xlsx`` and
selected with ``--input`` without changing the generated format.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency error is user-facing
    raise SystemExit("openpyxl が必要です。pip install openpyxl を実行してください。") from exc


ACTION_KEYS = {
    "普通攻撃間隔": "normalAttackInterval",
    "普通攻撃_基本": "basicAttack",
    "普通攻撃_強化": "enhancedAttack",
    "低学年": "lowSkill",
    "高学年": "highSkill",
}
DEFAULT_SHEETS = {
    "speed": "スキル速度",
    "timing": "スキルタイミング",
    "movement": "行動間移動",
    "object_base": "生成物基礎設定",
    "object_timing": "生成物タイミング",
    "end_conditions": "終了条件",
    "support_status": "対応状況",
}
SUPPORT_STATUS_HEADERS = ("id", "使徒名", "通常", "アサイド", "愛用品", "備考")
IMPLEMENTED_STATUSES = frozenset({"済", "暫定"})
KNOWN_IMPLEMENTATION_STATUSES = frozenset({"済", "暫定", "途中", "未"})
IMPLEMENTATION_COMPONENTS = (
    ("normal", "通常"),
    ("aside", "アサイド"),
    ("favorite", "愛用品"),
)
OWN_EFFECT_ACTION_MARKERS = {
    "basicAttack": "_basic_",
    "enhancedAttack": "_enhanced_",
    "lowSkill": "_low_",
    "highSkill": "_high_",
}
OWN_EFFECT_ID_RE = re.compile(r"^(?P<prefix>[A-Za-z0-9]+)_(?P<action>basic|enhanced|low|high)_e\d+$", re.IGNORECASE)
PLAYBACK_RATES = {
    "▶": 1.3,
    "▶▶": 1.69,
    "▶▶▶": 1.95,
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def optional_number(value: Any) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            return None
        return int(value) if float(value).is_integer() else float(value)
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def optional_bool(value: Any) -> bool | None:
    if value is None or clean(value) == "":
        return None
    if isinstance(value, bool):
        return value
    text = clean(value).lower()
    if text in {"true", "1", "yes", "on", "○", "あり", "対象"}:
        return True
    if text in {"false", "0", "no", "off", "×", "なし", "対象外"}:
        return False
    return None


def sheet_rows(workbook, sheet_name: str, required: bool = True) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        if required:
            raise ValueError(f"必要なシートがありません: {sheet_name}")
        return []
    worksheet = workbook[sheet_name]
    headers = [clean(cell.value) for cell in worksheet[1]]
    rows: list[dict[str, Any]] = []
    for line_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {header: value for header, value in zip(headers, values) if header}
        if not any(clean(value) for value in row.values()):
            continue
        row["__line__"] = line_number
        rows.append(row)
    return rows


def support_status_rows(workbook, sheet_name: str) -> dict[str, dict[str, Any]]:
    """Read per-component DPS implementation status from the source-of-truth sheet.

    Empty component cells deliberately mean ``未``.  This keeps a newly added
    column/row safe by default rather than accidentally enabling a configuration.
    """
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"必要なシートがありません: {sheet_name}")
    worksheet = workbook[sheet_name]
    headers = tuple(clean(cell.value) for cell in worksheet[1])
    missing = [header for header in SUPPORT_STATUS_HEADERS if header not in headers]
    if missing:
        raise ValueError(
            f"{sheet_name}: 必須ヘッダーがありません: {', '.join(missing)} "
            f"(検出: {', '.join(header or '(空)' for header in headers)})"
        )
    statuses: dict[str, dict[str, str]] = {}
    for row in sheet_rows(workbook, sheet_name):
        raw_id = clean(row.get("id"))
        if not raw_id:
            raise ValueError(f"{sheet_name} {row['__line__']}行: idが空です")
        key = raw_id.lower()
        component_statuses: dict[str, str] = {}
        for component_key, header in IMPLEMENTATION_COMPONENTS:
            status = clean(row.get(header)) or "未"
            if status not in KNOWN_IMPLEMENTATION_STATUSES:
                raise ValueError(
                    f"{sheet_name} {row['__line__']}行: {header}の未知の対応状況 {status!r} ({raw_id})"
                )
            component_statuses[component_key] = status
        if key in statuses:
            raise ValueError(
                f"{sheet_name} {row['__line__']}行: id {raw_id} が"
                f" {statuses[key]['sourceLine']}行と重複しています"
            )
        statuses[key] = {
            "id": key,
            "sourceId": raw_id,
            "name": clean(row.get("使徒名")),
            "statuses": component_statuses,
            "note": clean(row.get("備考")),
            "sourceLine": row["__line__"],
        }
    return statuses


def convert_to_game_frames(value: Any, unit: Any) -> float | int | None:
    number = optional_number(value)
    label = clean(unit)
    if number is None or not label:
        return None
    if label in {"ゲームF", "gameF"}:
        result = Decimal(str(number))
    elif label in {"ゲーム秒", "秒", "gameSecond"}:
        result = Decimal(str(number)) * Decimal("60")
    elif label.startswith("録画F@"):
        speed = label.removeprefix("録画F@").strip()
        rate = PLAYBACK_RATES.get(speed)
        if rate is None:
            return None
        result = Decimal(str(number)) * Decimal(str(rate))
    else:
        return None
    return int(result.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def source_time(value: Any, unit: Any) -> dict[str, Any]:
    return {
        "value": optional_number(value),
        "unit": clean(unit),
        "gameFrames": convert_to_game_frames(value, unit),
    }


def source_time_with_default(value: Any, unit: Any, default_unit: str) -> dict[str, Any]:
    """Resolve a time value while retaining whether the workbook omitted its unit."""
    has_value = optional_number(value) is not None
    source = source_time(value, clean(unit) or (default_unit if has_value else ""))
    if has_value and not clean(unit):
        source["unitInferred"] = True
    return source


def ensure_apostle(apostles: dict[str, dict], raw_id: Any, name: Any) -> dict:
    source_id = clean(raw_id)
    key = source_id.lower()
    if not key:
        raise ValueError("使徒idが空の行があります")
    entry = apostles.setdefault(
        key,
        {
            "id": key,
            "sourceId": source_id,
            "name": clean(name),
            "initialActionDelayFrames": 60,
            "initialActionDelaySource": "default",
            "measuredNormalAttackIntervalFrames": None,
            "normalAttackIntervalFrames": None,
            "normalAttackIntervalSource": None,
            "normalAttackIntervalVariants": [],
            "actions": {},
        },
    )
    incoming_name = clean(name)
    if incoming_name and entry["name"] and entry["name"] != incoming_name:
        raise ValueError(f"使徒名が一致しません: {source_id} ({entry['name']} / {incoming_name})")
    if incoming_name:
        entry["name"] = incoming_name
    return entry


def ensure_action(entry: dict, action_key: str, label: str) -> dict:
    return entry["actions"].setdefault(
        action_key,
        {
            "label": label,
            "branch": "",
            "motionFrames": None,
            "motionSource": None,
            "motionVariants": [],
            "researchStatus": "",
            "note": "",
            "timingEvents": [],
            "timingPatterns": [],
            "generatedObjects": [],
        },
    )


def normalize_cross_apostle_effect_id(
    raw_id: Any, action_key: str, effect_id: Any, warnings: list[str], source: str
) -> str:
    """Repair only an unmistakable copied direct-damage effect ID.

    The workbook remains untouched.  A timing row for Barong's enhanced attack
    cannot intentionally reference e.g. ``Renewa_enhanced_e01``: it would make
    the runtime resolve the wrong skill.  Keep generic artifacts/status IDs
    intact; this guard applies solely to action-shaped IDs with another apostle
    prefix and records the source row as a warning for workbook correction.
    """
    value = clean(effect_id)
    expected = clean(raw_id)
    marker = OWN_EFFECT_ACTION_MARKERS.get(action_key)
    if not value or not expected or not marker:
        return value
    matched = OWN_EFFECT_ID_RE.fullmatch(value)
    if not matched or matched.group("prefix").lower() == expected.lower():
        return value
    if f"_{matched.group('action').lower()}_" != marker:
        return value
    corrected = f"{expected}_{value[len(matched.group('prefix')) + 1:]}"
    warnings.append(f"{source}: 他使徒の効果ID {value} を {corrected} として補正しました")
    return corrected


def build_data(input_path: Path, sheet_names: dict[str, str]) -> dict:
    workbook = load_workbook(input_path, data_only=True, read_only=False)
    speed_rows = sheet_rows(workbook, sheet_names["speed"])
    timing_rows = sheet_rows(workbook, sheet_names["timing"])
    movement_sheet = sheet_names.get("movement", DEFAULT_SHEETS["movement"])
    movement_rows = sheet_rows(workbook, movement_sheet, required=False)
    object_base_rows = sheet_rows(workbook, sheet_names["object_base"], required=False)
    object_timing_rows = sheet_rows(workbook, sheet_names["object_timing"], required=False)
    end_condition_sheet = sheet_names.get("end_conditions", DEFAULT_SHEETS["end_conditions"])
    end_condition_rows = sheet_rows(workbook, end_condition_sheet, required=False)
    support_status_sheet = sheet_names.get("support_status", DEFAULT_SHEETS["support_status"])
    support_statuses = support_status_rows(workbook, support_status_sheet)
    apostles: dict[str, dict] = {}
    warnings: list[str] = []

    for row in speed_rows:
        line = row["__line__"]
        action_label = clean(row.get("動作名"))
        action_key = ACTION_KEYS.get(action_label)
        if not action_key:
            if action_label and action_label != "ターン回転":
                warnings.append(f"{sheet_names['speed']} {line}行: 未対応の動作名 {action_label}")
            continue
        entry = ensure_apostle(apostles, row.get("id"), row.get("使徒"))
        action = ensure_action(entry, action_key, action_label)
        incoming_branch = clean(row.get("分岐"))
        if not action["branch"]:
            action["branch"] = incoming_branch
        elif action["branch"] != incoming_branch:
            # Branched actions do not have one representative branch at the
            # action level. Keep branch information on each variant instead.
            action["branch"] = ""
        action["researchStatus"] = clean(row.get("調査状態"))
        action["note"] = clean(row.get("備考"))

        interval_source = source_time_with_default(
            row.get("普通攻撃間隔値"), row.get("普通攻撃間隔単位"), "録画F@▶"
        )
        if interval_source["value"] is not None:
            if interval_source["gameFrames"] is None and interval_source["unit"]:
                warnings.append(f"{sheet_names['speed']} {line}行: 普通攻撃間隔の単位を解決できません")
            elif interval_source["gameFrames"] is not None:
                variant = {
                    "actionKey": action_key,
                    "branch": clean(row.get("分岐")),
                    **interval_source,
                }
                if variant not in entry["normalAttackIntervalVariants"]:
                    entry["normalAttackIntervalVariants"].append(variant)
            if interval_source["gameFrames"] is not None and entry["normalAttackIntervalFrames"] is None:
                entry["measuredNormalAttackIntervalFrames"] = interval_source["value"]
                entry["normalAttackIntervalFrames"] = interval_source["gameFrames"]
                entry["normalAttackIntervalSource"] = interval_source

        motion_source = source_time_with_default(
            row.get("モーション値"), row.get("モーション単位"), "ゲームF"
        )
        if motion_source["value"] is not None and motion_source["gameFrames"] is None and motion_source["unit"]:
            warnings.append(f"{sheet_names['speed']} {line}行: モーションの単位を解決できません")
        if motion_source["gameFrames"] is not None:
            motion_variant = {
                "branch": clean(row.get("分岐")),
                **motion_source,
                "researchStatus": clean(row.get("調査状態")),
                "note": clean(row.get("備考")),
                "sourceLine": line,
            }
            if motion_variant not in action["motionVariants"]:
                action["motionVariants"].append(motion_variant)
            if not motion_variant["branch"] or action["motionFrames"] is None:
                action["motionFrames"] = motion_source["gameFrames"]
                action["motionSource"] = motion_source

    for row in timing_rows:
        line = row["__line__"]
        action_label = clean(row.get("動作名"))
        action_key = ACTION_KEYS.get(action_label)
        if not action_key:
            continue
        entry = ensure_apostle(apostles, row.get("id"), row.get("使徒"))
        action = ensure_action(entry, action_key, action_label)

        timing_mode = clean(row.get("発生方式"))
        adoption = clean(row.get("採用区分"))
        pattern_id = clean(row.get("パターンID"))
        is_pattern = timing_mode in {"初回+周期", "初回・周期", "初回+周期+終端", "パターン"} \
            or adoption == "採用"
        if is_pattern:
            initial_source = source_time(row.get("発生値"), row.get("発生単位"))
            interval_source = source_time(row.get("反復間隔値"), row.get("反復間隔単位"))
            if initial_source["value"] is None:
                warnings.append(f"{sheet_names['timing']} {line}行: パターンの初回発生値が空です")
            elif initial_source["gameFrames"] is None:
                warnings.append(f"{sheet_names['timing']} {line}行: パターンの初回発生単位を解決できません")
            if interval_source["value"] is None:
                warnings.append(f"{sheet_names['timing']} {line}行: パターンの反復間隔値が空です")
            elif interval_source["gameFrames"] is None:
                warnings.append(f"{sheet_names['timing']} {line}行: パターンの反復間隔単位を解決できません")
            action["timingPatterns"].append({
                "id": pattern_id or f"{entry['id']}:{action_key}:pattern:{line}",
                "branch": clean(row.get("分岐")),
                "order": optional_number(row.get("発生順")),
                "effectKind": clean(row.get("効果種別")),
                "effectId": normalize_cross_apostle_effect_id(
                    entry["sourceId"], action_key, row.get("効果ID") or row.get("effectId"), warnings,
                    f"{sheet_names['timing']} {line}行",
                ),
                "lv1PerHitMultiplier": optional_number(row.get("Lv1・1ヒット倍率(%)")),
                "initialFrame": initial_source["gameFrames"],
                "initialSourceTime": initial_source if initial_source["value"] is not None else None,
                "intervalFrames": interval_source["gameFrames"],
                "intervalSourceTime": interval_source if interval_source["value"] is not None else None,
                "timeOrigin": clean(row.get("時間基準")) or "動作開始",
                "countExpression": clean(row.get("発生回数式")),
                "endMode": clean(row.get("終了方式")) or "回数",
                "researchStatus": clean(row.get("発生調査状態")),
                "adoption": adoption or "採用",
                "note": clean(row.get("備考")),
                "sourceLine": line,
            })
            continue

        timing_source = source_time(row.get("発生値"), row.get("発生単位"))
        if timing_source["value"] is not None and timing_source["gameFrames"] is None and timing_source["unit"]:
            warnings.append(f"{sheet_names['timing']} {line}行: 発生値はありますが単位を解決できません")
        guarantee_source = source_time(row.get("保証値"), row.get("保証単位"))
        event = {
            "branch": clean(row.get("分岐")),
            "order": optional_number(row.get("発生順")),
            "effectKind": clean(row.get("効果種別")),
            "effectId": normalize_cross_apostle_effect_id(
                entry["sourceId"], action_key, row.get("効果ID") or row.get("effectId"), warnings,
                f"{sheet_names['timing']} {line}行",
            ),
            "lv1PerHitMultiplier": optional_number(row.get("Lv1・1ヒット倍率(%)")),
            "guaranteeFrames": guarantee_source["gameFrames"],
            "guaranteeSource": guarantee_source if guarantee_source["value"] is not None else None,
            "frame": timing_source["gameFrames"],
            "sourceTime": timing_source if timing_source["value"] is not None else None,
            "researchStatus": clean(row.get("発生調査状態")),
            "timingMode": timing_mode or "個別",
            "adoption": adoption or "",
            "note": clean(row.get("備考")),
            "sourceLine": line,
        }
        action["timingEvents"].append(event)

    movement_transition_count = 0
    movement_keys: set[tuple[str, str, str, str, str]] = set()
    for row in movement_rows:
        line = row["__line__"]
        entry = ensure_apostle(apostles, row.get("id"), row.get("使徒"))
        transition_id = clean(row.get("遷移ID") or row.get("移動ID"))
        from_label = clean(row.get("遷移元動作"))
        to_label = clean(row.get("遷移先動作"))
        from_action_key = ACTION_KEYS.get(from_label)
        to_action_key = ACTION_KEYS.get(to_label)
        if not from_action_key:
            warnings.append(f"{movement_sheet} {line}行: 未対応の遷移元動作 {from_label or '(空)'}")
            continue
        if not to_action_key:
            warnings.append(f"{movement_sheet} {line}行: 未対応の遷移先動作 {to_label or '(空)'}")
            continue
        movement_source = source_time(row.get("移動値"), row.get("移動単位"))
        if movement_source["value"] is None:
            warnings.append(f"{movement_sheet} {line}行: 移動値が空です")
            continue
        if movement_source["gameFrames"] is None:
            warnings.append(f"{movement_sheet} {line}行: 移動単位を解決できません")
            continue
        from_branch = clean(row.get("遷移元分岐"))
        to_branch = clean(row.get("遷移先分岐"))
        unique_key = (entry["id"], from_action_key, from_branch, to_action_key, to_branch)
        if unique_key in movement_keys:
            warnings.append(
                f"{movement_sheet} {line}行: {from_label}{('/' + from_branch) if from_branch else ''}"
                f" → {to_label}{('/' + to_branch) if to_branch else ''} が重複しています"
            )
            continue
        movement_keys.add(unique_key)
        entry.setdefault("movementTransitions", []).append({
            "id": transition_id or f"{entry['id']}:{from_action_key}:{from_branch}:{to_action_key}:{to_branch}",
            "fromActionKey": from_action_key,
            "fromBranch": from_branch,
            "toActionKey": to_action_key,
            "toBranch": to_branch,
            "frames": movement_source["gameFrames"],
            "sourceTime": movement_source,
            "researchStatus": clean(row.get("調査状態")),
            "note": clean(row.get("備考")),
            "sourceLine": line,
        })
        movement_transition_count += 1

    object_events: dict[str, list[dict]] = defaultdict(list)
    for row in object_timing_rows:
        line = row["__line__"]
        object_id = clean(row.get("生成物ID"))
        if not object_id:
            # Data-validation/default FALSE cells can extend far below the actual
            # table. They are formatting residue, not incomplete timing rows.
            has_content = any(
                clean(value)
                for key, value in row.items()
                if key not in {"__line__", "反復対象"}
            )
            if not has_content:
                continue
            warnings.append(f"{sheet_names['object_timing']} {line}行: 生成物IDが空です")
            continue
        timing_source = source_time(row.get("発生値"), row.get("発生単位"))
        if timing_source["value"] is not None and timing_source["gameFrames"] is None and timing_source["unit"]:
            warnings.append(f"{sheet_names['object_timing']} {line}行: 発生値はありますが単位を解決できません")
        repeat_target = optional_bool(row.get("反復対象"))
        if clean(row.get("反復対象")) and repeat_target is None:
            warnings.append(f"{sheet_names['object_timing']} {line}行: 反復対象を真偽値として解決できません")
        object_events[object_id].append({
            "branch": clean(row.get("分岐")),
            "instanceOrder": optional_number(row.get("個体順")),
            "eventType": clean(row.get("イベント種別")),
            "order": optional_number(row.get("イベント順")),
            "effectKind": clean(row.get("効果種別")),
            "effectId": clean(row.get("効果ID")),
            "lv1PerHitMultiplier": optional_number(row.get("Lv1・1ヒット倍率(%)")),
            "frame": timing_source["gameFrames"],
            "sourceTime": timing_source if timing_source["value"] is not None else None,
            "timeOrigin": clean(row.get("時間基準")),
            "repeatTarget": repeat_target,
            "researchStatus": clean(row.get("発生調査状態")),
            "note": clean(row.get("備考")),
            "sourceLine": line,
        })

    object_end_conditions: dict[str, list[dict]] = defaultdict(list)
    for row in end_condition_rows:
        line = row["__line__"]
        object_id = clean(row.get("生成物ID"))
        if not object_id:
            warnings.append(f"{end_condition_sheet} {line}行: 生成物IDが空です")
            continue
        condition_type = clean(row.get("終了条件"))
        if not condition_type:
            warnings.append(f"{end_condition_sheet} {line}行: 終了条件が空です")
        condition_value = optional_number(row.get("終了条件値"))
        condition_unit = clean(row.get("終了条件単位"))
        condition_frames = convert_to_game_frames(condition_value, condition_unit)
        is_time_condition = "時間" in condition_type or condition_unit in {
            "ゲームF", "gameF", "ゲーム秒", "秒", "gameSecond",
        } or condition_unit.startswith("録画F@")
        if condition_value is not None and is_time_condition and not condition_unit:
            warnings.append(f"{end_condition_sheet} {line}行: 時間条件の単位が空です")
        elif condition_value is not None and is_time_condition and condition_frames is None:
            warnings.append(f"{end_condition_sheet} {line}行: 時間条件の単位を解決できません")
        object_end_conditions[object_id].append({
            "order": optional_number(row.get("条件順")),
            "conditionType": condition_type,
            "conditionValue": condition_value,
            "conditionUnit": condition_unit,
            "conditionFrames": condition_frames,
            "join": clean(row.get("条件結合")),
            "endAction": clean(row.get("終了処理")),
            "researchStatus": clean(row.get("調査状態")),
            "note": clean(row.get("備考")),
            "sourceLine": line,
        })

    for conditions in object_end_conditions.values():
        conditions.sort(key=lambda item: (
            item["order"] is None, item["order"] or 0, item["sourceLine"],
        ))

    known_object_ids: set[str] = set()
    object_base_lines: dict[str, int] = {}
    for row in object_base_rows:
        line = row["__line__"]
        object_id = clean(row.get("生成物ID"))
        if not object_id:
            warnings.append(f"{sheet_names['object_base']} {line}行: 生成物IDが空です")
            continue
        if object_id in known_object_ids:
            warnings.append(
                f"{sheet_names['object_base']} {line}行: 生成物ID {object_id} が"
                f" {object_base_lines[object_id]}行と重複しています"
            )
        known_object_ids.add(object_id)
        object_base_lines.setdefault(object_id, line)
        action_label = clean(row.get("動作名"))
        action_key = ACTION_KEYS.get(action_label)
        if not action_key:
            warnings.append(f"{sheet_names['object_base']} {line}行: 未対応の動作名 {action_label}")
            continue
        entry = ensure_apostle(apostles, row.get("id"), row.get("使徒"))
        action = ensure_action(entry, action_key, action_label)
        spawn_source = source_time(row.get("生成発生値"), row.get("生成発生単位"))
        first_source = source_time(row.get("初回発生値"), row.get("初回発生単位"))
        interval_source = source_time(row.get("反復間隔値"), row.get("反復間隔単位"))
        for label, source in (
            ("生成発生", spawn_source),
            ("初回発生", first_source),
            ("反復間隔", interval_source),
        ):
            if source["value"] is not None and not source["unit"]:
                warnings.append(f"{sheet_names['object_base']} {line}行: {label}値がありますが単位が空です")
            elif source["value"] is not None and source["gameFrames"] is None:
                warnings.append(f"{sheet_names['object_base']} {line}行: {label}単位を解決できません")
        end_condition = clean(row.get("終了条件"))
        end_conditions = object_end_conditions.get(object_id, [])
        if end_condition == "終了条件シート参照" and not end_conditions:
            warnings.append(f"{sheet_names['object_base']} {line}行: {object_id} の終了条件行がありません")
        elif end_conditions and end_condition != "終了条件シート参照":
            warnings.append(
                f"{sheet_names['object_base']} {line}行: {object_id} に終了条件行がありますが"
                "終了条件シート参照になっていません"
            )
        timing_mode = clean(row.get("タイミング方式"))
        timing_events = object_events.get(object_id, [])
        if (
            "周期" in timing_mode
            and interval_source["gameFrames"] is not None
            and timing_events
            and not any(event.get("repeatTarget") is True for event in timing_events)
        ):
            warnings.append(
                f"{sheet_names['object_base']} {line}行: {object_id} は周期設定ですが"
                "反復対象のイベントがありません"
            )
        generated = {
            "id": object_id,
            "name": clean(row.get("生成物名")),
            "objectType": clean(row.get("生成物種別")),
            "branch": clean(row.get("分岐")),
            "spawnMode": clean(row.get("生成方式")),
            "spawnCount": optional_number(row.get("生成数")),
            "spawnCountEffectId": clean(row.get("生成数効果ID")),
            "spawnFrame": spawn_source["gameFrames"],
            "spawnSource": spawn_source if spawn_source["value"] is not None else None,
            "spawnOrigin": clean(row.get("生成基準")),
            "timingMode": timing_mode,
            "firstEventFrames": first_source["gameFrames"],
            "firstEventSource": first_source if first_source["value"] is not None else None,
            "repeatIntervalFrames": interval_source["gameFrames"],
            "repeatIntervalSource": interval_source if interval_source["value"] is not None else None,
            "repeatCount": optional_number(row.get("繰り返し回数")),
            "endCondition": end_condition,
            "endConditionValue": optional_number(row.get("終了条件値")),
            "endConditions": end_conditions,
            "respawnPolicy": clean(row.get("再生成時処理")),
            "maxInstances": optional_number(row.get("最大存在数")),
            "statReference": clean(row.get("ステータス参照")),
            "attackSpeedReference": clean(row.get("攻撃速度参照")),
            "attackSpeedScope": clean(
                row.get("攻撃速度適用範囲") or row.get("速度適用範囲")
            ),
            "attackSpeedChangePolicy": clean(
                row.get("攻撃速度変更反映") or row.get("速度変更反映")
            ),
            "cancelPolicy": clean(row.get("キャンセル後処理")),
            "researchStatus": clean(row.get("調査状態")),
            "note": clean(row.get("備考")),
            "timingEvents": timing_events,
            "sourceLine": line,
        }
        action["generatedObjects"].append(generated)

    for object_id in sorted(set(object_events) - known_object_ids):
        warnings.append(f"{sheet_names['object_timing']}: 基礎設定がない生成物ID {object_id}")
    for object_id in sorted(set(object_end_conditions) - known_object_ids):
        warnings.append(f"{end_condition_sheet}: 基礎設定がない生成物ID {object_id}")

    for entry in apostles.values():
        if entry.get("movementTransitions"):
            entry["movementTransitions"].sort(key=lambda item: (
                item["fromActionKey"], item["fromBranch"], item["toActionKey"],
                item["toBranch"], item["sourceLine"],
            ))
        for action in entry["actions"].values():
            action["motionVariants"].sort(key=lambda item: (
                item["branch"], item["sourceLine"],
            ))
            action["timingPatterns"].sort(key=lambda item: (
                item["branch"], item["order"] is None, item["order"] or 0,
                item["sourceLine"],
            ))
            action["timingEvents"].sort(key=lambda item: (
                item["branch"], item["order"] is None, item["order"] or 0,
                item["frame"] is None, item["frame"] or 0,
            ))
            for generated in action["generatedObjects"]:
                generated["timingEvents"].sort(key=lambda item: (
                    item["branch"], item["instanceOrder"] is None, item["instanceOrder"] or 0,
                    item["order"] is None, item["order"] or 0,
                ))

    source_apostle_count = len(apostles)
    usable_apostles: dict[str, dict] = {}
    incomplete_apostles: list[dict[str, Any]] = []
    for key, entry in apostles.items():
        support_status = support_statuses.get(key)
        if support_status:
            entry["implementationStatuses"] = dict(support_status["statuses"])
            entry["implementationNote"] = support_status["note"]
            if support_status["name"] and entry["name"] and support_status["name"] != entry["name"]:
                warnings.append(
                    f"{support_status_sheet} {support_status['sourceLine']}行: {support_status['sourceId']} の使徒名が"
                    f"タイミングデータと一致しません ({support_status['name']} / {entry['name']})"
                )
        else:
            # Keep executable records structurally uniform.  The registry still
            # requires an actual support-status row, so this default cannot make
            # an unregistered apostle selectable.
            entry["implementationStatuses"] = {
                component_key: "未" for component_key, _ in IMPLEMENTATION_COMPONENTS
            }
            entry["implementationNote"] = ""
        missing: list[str] = []
        if entry["normalAttackIntervalFrames"] is None:
            missing.append("普通攻撃間隔")
        if entry["actions"].get("basicAttack", {}).get("motionFrames") is None:
            missing.append("基本攻撃モーション")
        if missing:
            incomplete_apostles.append({
                "id": key,
                "name": entry["name"],
                "missing": missing,
            })
        else:
            usable_apostles[key] = entry
    return {
        "version": 8,
        "source": {
            "workbook": input_path.name,
            "sheets": sheet_names,
        },
        "timeBase": {
            "framesPerGameSecond": 60,
            "subframeTicks": 10,
            "playbackRates": PLAYBACK_RATES,
            "convertedFrameRounding": "halfUp",
            "defaultInitialActionDelayFrames": 60,
        },
        "summary": {
            "apostles": source_apostle_count,
            "usableApostles": len(usable_apostles),
            "generatedObjects": len(known_object_ids),
            "endConditions": sum(len(items) for items in object_end_conditions.values()),
            "movementTransitions": movement_transition_count,
            "timingPatterns": sum(
                len(action.get("timingPatterns", []))
                for entry in usable_apostles.values()
                for action in entry.get("actions", {}).values()
            ),
            "implementationStatuses": {
                component_key: {
                    status: sum(
                        1 for item in support_statuses.values()
                        if item["statuses"][component_key] == status
                    )
                    for status in sorted(KNOWN_IMPLEMENTATION_STATUSES)
                }
                for component_key, _ in IMPLEMENTATION_COMPONENTS
            },
        },
        "apostles": usable_apostles,
        "supportStatuses": support_statuses,
        "incompleteApostles": incomplete_apostles,
        "warnings": warnings,
    }


def write_javascript(output_path: Path, data: dict) -> None:
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        "// Generated by tools/generate-dps-timing-data.py. Do not edit directly.\n"
        f"const DPS_TIMING_DATA = {payload};\n"
        "if (typeof globalThis !== 'undefined') globalThis.DPS_TIMING_DATA = DPS_TIMING_DATA;\n"
        "if (typeof module !== 'undefined' && module.exports) module.exports = DPS_TIMING_DATA;\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="skillmotionシートからDPS用データを生成します")
    parser.add_argument("--input", type=Path, default=Path(__file__).with_name("trickcal_skillmotion.xlsx"))
    parser.add_argument("--output", type=Path, default=Path(__file__).resolve().parent.parent / "dps-timing-data.js")
    parser.add_argument("--speed-sheet", default=DEFAULT_SHEETS["speed"])
    parser.add_argument("--timing-sheet", default=DEFAULT_SHEETS["timing"])
    parser.add_argument("--movement-sheet", default=DEFAULT_SHEETS["movement"])
    parser.add_argument("--object-base-sheet", default=DEFAULT_SHEETS["object_base"])
    parser.add_argument("--object-timing-sheet", default=DEFAULT_SHEETS["object_timing"])
    parser.add_argument("--end-condition-sheet", default=DEFAULT_SHEETS["end_conditions"])
    parser.add_argument("--support-status-sheet", default=DEFAULT_SHEETS["support_status"])
    parser.add_argument("--strict", action="store_true", help="警告が1件でもあれば終了コード2にする")
    args = parser.parse_args()
    sheet_names = {
        "speed": args.speed_sheet,
        "timing": args.timing_sheet,
        "movement": args.movement_sheet,
        "object_base": args.object_base_sheet,
        "object_timing": args.object_timing_sheet,
        "end_conditions": args.end_condition_sheet,
        "support_status": args.support_status_sheet,
    }
    input_path = args.input.resolve()
    if not input_path.exists():
        raise SystemExit(f"入力ファイルがありません: {input_path}")
    data = build_data(input_path, sheet_names)
    write_javascript(args.output.resolve(), data)
    print(
        f"Generated {args.output.resolve()} "
        f"({data['summary']['apostles']} apostles, "
        f"{data['summary']['usableApostles']} usable, "
        f"{data['summary']['generatedObjects']} generated objects, "
        f"{data['summary']['endConditions']} end conditions, "
        f"{data['summary']['movementTransitions']} movement transitions, "
        f"{data['summary']['timingPatterns']} timing patterns, "
        f"{len(data['warnings'])} warnings)"
    )
    for warning in data["warnings"]:
        print(f"WARNING: {warning}")
    if args.strict and data["warnings"]:
        sys.exit(2)


if __name__ == "__main__":
    main()
